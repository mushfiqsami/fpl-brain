"""Output: a readable console briefing and a refreshed Excel workbook."""
from __future__ import annotations
import os
from .model import POS_NAME

W = 96


def rule(ch="="):
    return ch * W


def header(t):
    return f"\n{rule()}\n {t}\n{rule()}"


def briefing(ctx):
    """ctx is assembled by run.py. Returns the full text report."""
    L = []
    a = L.append
    a(header(f"FPL BRAIN  —  Gameweek {ctx['gw']} briefing"))
    a(f" Generated {ctx['generated']}   |   deadline {ctx.get('deadline','?')}")
    a(f" Model state: {ctx['teams_with_data']} clubs with {ctx['games_played']} games of "
      f"2026/27 data | prior weight {ctx['prior_share']:.0%}")
    cal = ctx.get("calibration", {})
    if cal:
        a(" Calibration: " + "  ".join(f"{k} x{v:.3f}" for k, v in cal.items()))

    if ctx.get("squad_source") == "live":
        a(header("YOUR SQUAD"))
        a(f" Team: {ctx.get('entry_name','?')}   Bank £{ctx['bank']:.1f}m   "
          f"Squad value £{ctx['squad_value']:.1f}m   Free transfers {ctx['free_transfers']}")
    else:
        a(header("NO TEAM ID SET — showing the optimal squad from scratch"))
        a(" Add your entry_id to config.json to get personalised transfer advice.")

    # ---- transfer recommendation ----
    a(header("RECOMMENDED ACTION"))
    p = ctx.get("plan")
    if p is None:
        a(" (no live squad to act on)")
    elif not p["out"]:
        a(" HOLD. No transfer clears its cost this week.")
        a(f" Best available swap was worth {ctx.get('best_swap_gain', 0):+.2f} pts over the "
          f"horizon, below the {ctx['ft_value']:.1f} pt value of banking the transfer.")
    else:
        for o, i in zip(p["out"], p["inn"]):
            a(f" OUT  {o['name']:20} {POS_NAME[o['pos']]:4} £{o['sell']:.1f}m")
            a(f" IN   {i['name']:20} {POS_NAME[i['pos']]:4} £{i['price']:.1f}m")
        a(f" Transfers: {len(p['out'])}   free used: {p['used_ft']}   hits: {p['hits']} "
          f"({-4*p['hits']} pts)")

    # ---- XI ----
    a(header(f"STARTING XI FOR GW{ctx['gw']}"))
    a(f" {'':3}{'Player':22}{'Club':6}{'Pos':5}{'£m':>6}{'EP':>7}{'p(play)':>9}  Fixture")
    for row in ctx["xi_rows"]:
        tag = "(C)" if row["is_cap"] else ("(V)" if row["is_vice"] else "")
        a(f" {tag:3}{row['name']:22}{row['club']:6}{row['pos']:5}{row['price']:6.1f}"
          f"{row['ep']:7.2f}{row['p_appear']:9.0%}  {row['fixture']}")
    a(f" {'':3}{'PROJECTED XI TOTAL':22}{'':11}{'':6}{ctx['xi_total']:7.2f}  "
      f"(captain doubled: {ctx['xi_total_c']:.2f})")

    a(header("BENCH  (in autosub order)"))
    for row in ctx["bench_rows"]:
        a(f" {row['order']}. {row['name']:22}{row['club']:6}{row['pos']:5}{row['price']:6.1f}"
          f"{row['ep']:7.2f}{row['p_appear']:9.0%}  {row['fixture']}")

    # ---- watchlist ----
    a(header("TRANSFER TARGETS  —  best expected points you do not own"))
    a(f" {'Player':22}{'Club':6}{'Pos':5}{'£m':>6}{'EP now':>8}{'EP next 5':>11}{'per £m':>9}  Note")
    for r in ctx["targets"][:15]:
        a(f" {r['name']:22}{r['club']:6}{r['pos']:5}{r['price']:6.1f}{r['ep']:8.2f}"
          f"{r['ep5']:11.2f}{r['value']:9.2f}  {r['note']}")

    # ---- sell candidates ----
    if ctx.get("sell_rows"):
        a(header("WEAKEST LINKS IN YOUR SQUAD"))
        for r in ctx["sell_rows"][:6]:
            a(f" {r['name']:22}{r['club']:6}{r['pos']:5}{r['price']:6.1f}{r['ep']:8.2f}"
              f"{r['ep5']:11.2f}   {r['note']}")

    # ---- fixture outlook ----
    a(header("FIXTURE OUTLOOK  —  next 5 gameweeks, by projected team xG"))
    a(f" {'Club':16}" + "".join(f"GW{g:<7}" for g in ctx["horizon"]) + "  5-GW total")
    for r in ctx["fixture_rows"]:
        cells = "".join(f"{c:<9}" for c in r["cells"])
        a(f" {r['club']:16}{cells}  {r['total']:.2f}")

    # ---- chips ----
    if ctx.get("chip_notes"):
        a(header("CHIP WATCH"))
        for n in ctx["chip_notes"]:
            a(" - " + n)

    a(header("HEALTH / CAVEATS"))
    for n in ctx["caveats"]:
        a(" - " + n)
    a("")
    return "\n".join(L)


def write_excel(ctx, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule

    F = "Arial"
    H1 = Font(name=F, size=13, bold=True, color="FFFFFF")
    H2 = Font(name=F, size=10, bold=True, color="FFFFFF")
    N = Font(name=F, size=10)
    B = Font(name=F, size=10, bold=True)
    S = Font(name=F, size=9, color="595959")
    NAVY = PatternFill("solid", fgColor="1F3864")
    MIDF = PatternFill("solid", fgColor="2E5C8A")
    thin = Side(style="thin", color="BFBFBF")
    BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    def sheet(name, title, cols, widths):
        ws = wb.create_sheet(name) if wb.sheetnames != ["Sheet"] else wb.active
        if ws.title == "Sheet":
            ws.title = name
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        c = ws.cell(1, 1, title); c.font = H1; c.fill = NAVY
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 24
        for i, (h, w) in enumerate(zip(cols, widths), start=1):
            cc = ws.cell(3, i, h); cc.font = H2; cc.fill = MIDF; cc.border = BOX
            cc.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A4"
        return ws

    def fill(ws, rows, start=4, fmts=None):
        for r, row in enumerate(rows, start=start):
            for cnum, v in enumerate(row, start=1):
                cell = ws.cell(r, cnum, v)
                cell.font = N; cell.border = BOX
                if fmts and cnum in fmts:
                    cell.number_format = fmts[cnum]
                if isinstance(v, (int, float)):
                    cell.alignment = Alignment(horizontal="center")
        return start + len(rows) - 1

    # Squad
    ws = sheet("Squad", f"GW{ctx['gw']} squad and starting XI",
               ["Role", "Player", "Club", "Pos", "£m", "EP", "p(play)", "Fixture", "Note"],
               [8, 22, 8, 6, 8, 8, 9, 26, 40])
    rows = [["XI" + (" (C)" if r["is_cap"] else " (V)" if r["is_vice"] else ""),
             r["name"], r["club"], r["pos"], r["price"], round(r["ep"], 2),
             r["p_appear"], r["fixture"], r.get("note", "")] for r in ctx["xi_rows"]]
    rows += [[f"Bench {r['order']}", r["name"], r["club"], r["pos"], r["price"],
              round(r["ep"], 2), r["p_appear"], r["fixture"], r.get("note", "")]
             for r in ctx["bench_rows"]]
    last = fill(ws, rows, fmts={5: "0.0", 6: "0.00", 7: "0%"})
    ws.conditional_formatting.add(f"F4:F{last}",
        ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile",
                       mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))

    # Targets
    ws = sheet("Targets", "Transfer targets ranked by expected points",
               ["Player", "Club", "Pos", "£m", "EP this GW", "EP next 5", "EP per £m",
                "p(play)", "Owned %", "Note"],
               [22, 8, 6, 8, 12, 12, 11, 9, 9, 40])
    last = fill(ws, [[r["name"], r["club"], r["pos"], r["price"], round(r["ep"], 2),
                      round(r["ep5"], 2), round(r["value"], 2), r["p_appear"],
                      r.get("owned", 0), r["note"]] for r in ctx["targets"]],
                fmts={4: "0.0", 5: "0.00", 6: "0.00", 7: "0.00", 8: "0%", 9: "0.0"})
    ws.auto_filter.ref = f"A3:J{last}"

    # Fixtures
    ws = sheet("Fixtures", "Projected team xG by gameweek",
               ["Club"] + [f"GW{g}" for g in ctx["horizon"]] + ["5-GW total"],
               [18] + [12] * len(ctx["horizon"]) + [12])
    last = fill(ws, [[r["club"]] + r["xgs"] + [round(r["total"], 2)] for r in ctx["fixture_rows"]],
                fmts={i: "0.00" for i in range(2, len(ctx["horizon"]) + 3)})
    col_end = chr(64 + len(ctx["horizon"]) + 1)
    ws.conditional_formatting.add(f"B4:{col_end}{last}",
        ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile",
                       mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))

    # Team ratings
    ws = sheet("TeamRatings", "Team strength — prior blended with 2026/27 observed xG",
               ["Club", "Games", "Attack", "Defence", "Observed xGF/gm", "Observed xGA/gm",
                "Prior weight"], [20, 8, 10, 10, 16, 16, 12])
    fill(ws, ctx["team_rating_rows"], fmts={3: "0.000", 4: "0.000", 5: "0.00", 6: "0.00", 7: "0%"})

    # Calibration
    ws = sheet("Calibration", "Model accuracy log — updated by `python run.py calibrate`",
               ["GW", "Position", "n", "MAE", "Bias", "Spearman", "Multiplier"],
               [8, 12, 8, 10, 10, 12, 12])
    fill(ws, ctx.get("calibration_rows", []), fmts={4: "0.000", 5: "0.000", 6: "0.000", 7: "0.000"})

    # Log
    ws = sheet("RunLog", "What this run did", ["Item", "Value"], [40, 90])
    fill(ws, ctx["log_rows"])

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path
